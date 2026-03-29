from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


INVENTORY_COLUMNS = [
    "item_id",
    "source_image",
    "photo_id",
    "file_name",
    "file_path",
    "detected_count_on_image",
    "title",
    "description",
    "dimensions_raw",
    "width",
    "height",
    "depth",
    "dimension_unit",
    "label_text",
    "detected_metadata",
    "notes",
    "confidence",
    "status",
    "created_at",
    "image_width_px",
    "image_height_px",
    "file_size",
    "file_created_at",
    "file_modified_at",
    "exif_date",
    "hash",
    "perceptual_hash",
    "duplicate_group",
    "duplicate_suspected",
    "category",
    "tag_1",
    "tag_2",
    "tag_3",
    "manual_review_required",
    "ocr_used",
    "ocr_text_raw",
    "hyperlink",
]


def build_summary(records: list[dict[str, object]], file_count: int) -> pd.DataFrame:
    dataframe = pd.DataFrame(records)
    metrics = [
        ("files_processed", file_count),
        ("rows_exported", len(records)),
        ("rows_with_dimensions", int(dataframe["dimensions_raw"].astype(bool).sum()) if not dataframe.empty else 0),
        ("rows_with_label_text", int(dataframe["label_text"].astype(bool).sum()) if not dataframe.empty else 0),
        ("rows_needing_manual_review", int(dataframe["manual_review_required"].astype(bool).sum()) if not dataframe.empty else 0),
        ("duplicate_suspected_rows", int(dataframe["duplicate_suspected"].astype(bool).sum()) if not dataframe.empty else 0),
    ]
    return pd.DataFrame(metrics, columns=["metric", "value"])


def export_csv(dataframe: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_path, index=False)


def export_xlsx(dataframe: pd.DataFrame, summary: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="inventory")
        summary.to_excel(writer, index=False, sheet_name="summary")

    workbook = load_workbook(output_path)
    inventory_sheet = workbook["inventory"]
    inventory_sheet.freeze_panes = "A2"
    inventory_sheet.auto_filter.ref = inventory_sheet.dimensions

    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in inventory_sheet[1]:
        cell.font = header_font

    for row in inventory_sheet.iter_rows():
        for cell in row:
            cell.alignment = top_alignment

    for index, _ in enumerate(INVENTORY_COLUMNS, start=1):
        letter = get_column_letter(index)
        max_length = 0
        for cell in inventory_sheet[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        inventory_sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)

    hyperlink_column = INVENTORY_COLUMNS.index("hyperlink") + 1
    source_column = INVENTORY_COLUMNS.index("source_image") + 1
    for row_index in range(2, inventory_sheet.max_row + 1):
        file_path = inventory_sheet.cell(row=row_index, column=hyperlink_column).value
        source_cell = inventory_sheet.cell(row=row_index, column=source_column)
        if file_path:
            source_cell.hyperlink = file_path
            source_cell.style = "Hyperlink"

    summary_sheet = workbook["summary"]
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions
    for cell in summary_sheet[1]:
        cell.font = header_font

    workbook.save(output_path)
